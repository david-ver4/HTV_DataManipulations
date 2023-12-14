import pandas as pd
import numpy as np
import openpyxl

# Load the input Excel file
input_file = "dateShiftedData3.xlsx"
input_workbook = pd.ExcelFile(input_file)

# Create an empty dictionary to store output DataFrames for each sheet
output_dfs = {}
constant_values = [
    "Weekday 1", "Weekend 1", "Weekday 2", "Weekend 2", "Weekday 3", "Weekend 3",
    "Weekday 4", "Weekend 4", "Weekday 5", "Weekend 5", "Weekday 6", "Weekend 6",
    "Weekday 7", "Weekend 7", "Weekday 8", "Weekend 8", "Weekday 9", "Weekend 9",
    "Weekday 10", "Weekend 10", "Weekday 11", "Weekend 11"
]
# Iterate over each sheet in the input workbook
for sheet_name in input_workbook.sheet_names:
    df = pd.read_excel(input_workbook, sheet_name, index_col=0)

    # Create an empty DataFrame for the output
    output_df = pd.DataFrame(index=["Weekday {}".format(i) for i in range(1, (len(df.columns) * 2) + 1)])
    # Set the first column to constant_values by repeating it
    # Iterate over each column in the input DataFrame
    for col in df.columns:
        data = df[col].fillna(0)  # Replace NaN values with 0

        # Calculate the number of averages to be taken
        num_averages = (len(data) - 13) // 7
        if (len(data) - 13) % 7 != 0:
            num_averages += 1

        idx = 11  # makes friday and saturdays the weekend for sleep data
        #idx = 12

        # Calculate and append the averages to the output DataFrame
        for i in range(num_averages):
            if (i + 1) % 2 == 1:  # If i+1 is odd, take the average of 5 consecutive cells
                start_idx = idx
                end_idx = idx + 5
                idx += 5
            else:  # If i+1 is even, take the average of 2 consecutive cells
                start_idx = idx
                end_idx = idx + 2
                idx += 2

            non_zero_values = data.iloc[start_idx:end_idx][data.iloc[start_idx:end_idx] != 0]
            avg = non_zero_values.mean() if len(non_zero_values) > 0 else 0

            # Create a new DataFrame with the results
            result_df = pd.DataFrame({f"{col}": [avg]}, index=[f"Weekday {i + 1}"])

            # Concatenate the new DataFrame with the output DataFrame
            output_df.loc[f"Weekday {i + 1}", col] = avg
            # Add the output DataFrame to the dictionary
    output_dfs[sheet_name] = output_df


def create_weekday_weekend_column(sheet):
    # Create the desired column values
    column_values = [
        "Weekday 1", "Weekend 1", "Weekday 2", "Weekend 2", "Weekday 3", "Weekend 3",
        "Weekday 4", "Weekend 4", "Weekday 5", "Weekend 5", "Weekday 6", "Weekend 6",
        "Weekday 7", "Weekend 7", "Weekday 8", "Weekend 8", "Weekday 9", "Weekend 9",
        "Weekday 10", "Weekend 10", "Weekday 11", "Weekend 11"
    ]

    # Iterate through the values and write them to the first column (column A)
    for index, value in enumerate(column_values, start=1):
        sheet.cell(row=index + 1, column=1, value=value)

    # Set row height for better readability
    for row in sheet.iter_rows(min_row=1, max_row=len(column_values) + 1, max_col=1):
        for cell in row:
            sheet.row_dimensions[cell.row].height = 20  # Adjust this value based on your preference

    # Stretch the width of the first column
    sheet.column_dimensions['A'].width = 25  # Adjust this value based on your preference

    # Clear cells in rows 24 and below
    for row in range(24, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = None

# Save the output DataFrames to a new Excel file with the same sheet names
output_file = "sleepDataWeekAverages.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name, output_df in output_dfs.items():
        sheet = input_workbook.parse(sheet_name)  # Access the sheet using parse
        sheet.sheet_state = 'visible'  # Set the sheet state to visible
        output_df.to_excel(writer, sheet_name=sheet_name)
input_file2 = output_file

    # Load the input workbook
wb = openpyxl.load_workbook(input_file2)

    # Iterate through each sheet in the workbook
for sheet_name in wb.sheetnames:
 sheet = wb[sheet_name]

        # Create the weekday-weekend column, stretch the first column, and clear cells in rows 24 and below for each sheet
 create_weekday_weekend_column(sheet)

    # Save the modified workbook to a new file
 wb.save(output_file)
