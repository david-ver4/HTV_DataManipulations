import openpyxl
from openpyxl.styles import PatternFill
import datetime

# Load the Excel file
workbook = openpyxl.load_workbook('dateShiftedData.xlsx')
worksheet = workbook.active  # You may need to specify the sheet name if it's not the active sheet

# Define a pattern fill for highlighting (e.g., yellow)
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Iterate through the rows, starting from the second row (assuming header is in the first row)
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    date_cell = row[0]
    date_value = date_cell.value

    if isinstance(date_value, datetime.datetime):
        # Check if the date is a weekend (Saturday or Sunday)
        if date_value.weekday() in [5, 6]:
            # If it's a weekend date, highlight the entire row
            for cell in row:
                cell.fill = highlight_fill

# Save the modified Excel file
workbook.save('dateShiftedData_highlighted_rows.xlsx')

