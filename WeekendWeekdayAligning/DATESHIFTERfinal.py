import pandas as pd
import openpyxl

# Load the consent dates data
consent_dates_df = pd.read_excel('Sp23_consent_dates.xlsx')

# Load the test6 data
workbook = openpyxl.load_workbook('SP23_sleepActivity.xlsx')

# Create a dictionary to map participant IDs to their consent dates
consent_date_dict = dict(zip(consent_dates_df['ID'], consent_dates_df['Consent Date']))

# Define the reference date (2/15/2023)
reference_date = pd.to_datetime('2/15/2023', format='%m/%d/%Y')

# Iterate through all sheets in the workbook
for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]

    # Initialize the row index (starts with row 2) and column index (starts with column B)
    row_index = 1  # Start from row 2
    column_index = 2  # Start from column B
    days_difference_dict = {}

    # Iterate through the rows of test6 and print participant ID and days difference
    for row in range(0, len(consent_dates_df)):
        participant_id = consent_dates_df.iloc[row, 0]
        consent_date = consent_date_dict.get(participant_id, '2/15/2023')
        consent_date = pd.to_datetime(consent_date, format='%m/%d/%Y')
        days_difference = (consent_date - reference_date).days

        days_difference_dict[participant_id] = days_difference

    # Iterate through the rows and insert blank cells using the days_difference_dict
    for participant_id, days_difference in days_difference_dict.items():
        if days_difference > 0:
            for i in range(days_difference):
                # Shift the cells down
                for row in range(worksheet.max_row, row_index, -1):
                    cell = worksheet.cell(row=row, column=column_index)
                    cell_offset = worksheet.cell(row=row + 1, column=column_index)
                    cell_offset.value = cell.value

            # Insert blank cells under the participant ID cell
            for i in range(1, days_difference + 1):
                cell = worksheet.cell(row=row_index + i, column=column_index)
                cell.value = ''

        # Increment the column index for the next participant
        column_index += 1

# Save the modified Excel file
workbook.save('dateShiftedData3.xlsx')

