import openpyxl

# Load the source Excel file with zero allocated for 10 hour threshold
source_file = '10hour_F23_weartimeActivity.xlsx'
source_sheet_name = 'TotalMinutesWearTime'
source_wb = openpyxl.load_workbook(source_file)
source_sheet = source_wb[source_sheet_name]

# Load the target Excel file that you want to allocate zeros for
target_file = 'Fall23_physicalActivity-10houred.xlsx'
target_wb = openpyxl.load_workbook(target_file)

# Iterate through each sheet in the target workbook
for target_sheet in target_wb.sheetnames:
    target_sheet = target_wb[target_sheet]

    # Iterate through cells in the source sheet
    for row_index, row in enumerate(source_sheet.iter_rows(min_row=2, min_col=2, values_only=True), start=2):
        for col_index, cell_value in enumerate(row, start=2):
            if cell_value == 0:
                # Set the corresponding cell in the target sheet to 0
                target_sheet.cell(row=row_index, column=col_index, value=0)

# Save the changes to the target file
target_wb.save(target_file)

# Close both workbooks
source_wb.close()
target_wb.close()



